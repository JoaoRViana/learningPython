import  customtkinter as ctk
from tkinter import *
from tkinter import messagebox
import openpyxl, xlrd
import pathlib
from openpyxl import Workbook

ctk.set_appearance_mode("System");
ctk.set_default_color_theme("blue");

class App(ctk.CTk):
    def __init__(self):     
        super().__init__()
        self.layout_config()
        self.appearance()
        self.system_title()
        self.system_sub_title()
        self.system_main()

    archive = pathlib.Path("Clients.xlsx")
    if archive.exists():
        pass
    else:
        archive =Workbook();
        paper = archive.active
        paper['A1']="Nome"
        paper['B1']="Telefone"
        paper['C1']="Endereço"
        paper['D1']='Idade'
        paper['E1']="Genêro"
        paper['F1']="Observações"

        archive.save("Clientes.xlsx")

    def layout_config(self):
        self.title("Sistema de Gestão de Clientes")
        self.geometry("700x500")

    def appearance(self):
        self.lb_apm = ctk.CTkLabel(self,text='Tema', bg_color='transparent',text_color=['#000','#fff']).place(x=50,y=430)
        self.opt_apm= ctk.CTkOptionMenu(self,values=['Light','Dark','System'],command=self.change_apm).place(x=50,y=460)

    def change_apm(self,new_appearance_mode):
        ctk.set_appearance_mode(new_appearance_mode)

    def system_title(self):
        frame = ctk.CTkFrame(self,width=600,height=50,corner_radius=0,bg_color=["#271",'#121'],fg_color=['#271','#121']).place(x=50,y=12)
        title = ctk.CTkLabel(frame, text='Sistema de Gestão de Clientes', bg_color=['#271','#121'],fg_color=['#271','#121'],font=('Century Gothic bold', 24), text_color=['#000','#fff'])
        title.pack(pady=20)

    def system_sub_title(self):
        span = ctk.CTkLabel(self,text='Por favor, preencha todos os campos do formulário!',font=('Century Gothic bold',16),text_color=['#000','#fff']).pack(pady=20)


    def system_main(self):

        def clear():
            name_value.set("");
            contact_value.set("")
            address_value.set("")
            age_value.set("")
            obs_text_box.delete(0.0,END)

        def submit():
            
            name = name_value.get();
            contact= contact_value.get()
            address = address_value.get()
            age = age_value.get()
            gender = gender_box.get()
            obs= obs_text_box.get(0.0,END)
            
            try:
                age_number=int(age);
                contact_number=int(contact)
                pass
            except :
                messagebox.showinfo("Sistema","Preenchimento incorreto da idade ou telefone")
                return

            if(len(name)<3 or age_number<1 or len(contact)<8 or len(address)<3 or len(obs)<3):
                messagebox.showinfo("Sistema","Há informações faltando")
                return

            archive = openpyxl.load_workbook("Clientes.xlsx")
            paper = archive.active
            paper.cell(column=1,row=paper.max_row+1,value=name)
            paper.cell(column=2,row=paper.max_row,value=contact)
            paper.cell(column=3,row=paper.max_row,value=address)
            paper.cell(column=4,row=paper.max_row,value=age)    
            paper.cell(column=5,row=paper.max_row,value=gender)
            paper.cell(column=6,row=paper.max_row,value=obs)
            
            archive.save(r"Clientes.xlsx")
            messagebox.showinfo("Sistema","Dados salvos com sucesso!")

            clear()

        
        name_value = StringVar()
        lb_name = ctk.CTkLabel(self,text='Nome:',font=('Century Gothic bold',16),text_color=['#000','#fff']).place(x=50,y=120)
        name_entry = ctk.CTkEntry(self,width=350,font=("Century Gohtic bold",16),textvariable=name_value,bg_color='transparent',fg_color='transparent').place(x=50,y=150)

        contact_value = StringVar()
        lb_contact = ctk.CTkLabel(self,text='Telefone:',font=('Century Gothic bold',16),text_color=['#000','#fff']).place(x=420,y=120)
        contact_entry = ctk.CTkEntry(self,width=200,font=("Century Gohtic bold",16),textvariable=contact_value,bg_color='transparent',fg_color='transparent').place(x=420,y=150)

        address_value = StringVar()
        lb_address = ctk.CTkLabel(self,text='Endereço:',font=('Century Gothic bold',16),text_color=['#000','#fff']).place(x=50,y=180)
        address_entry = ctk.CTkEntry(self,width=200,font=("Century Gohtic bold",16),textvariable=address_value,bg_color='transparent',fg_color='transparent').place(x=50,y=210)

        age_value = StringVar()
        lb_age_ = ctk.CTkLabel(self,text='Idade:',font=('Century Gothic bold',16),text_color=['#000','#fff']).place(x=260,y=180)
        age_entry = ctk.CTkEntry(self,width=150,font=("Century Gohtic bold",16),textvariable=age_value,bg_color='transparent',fg_color='transparent').place(x=260,y=210)

        lb_gender = ctk.CTkLabel(self,text='Gênero:',font=('Century Gothic bold',16),text_color=['#000','#fff']).place(x=420,y=180)
        gender_box = ctk.CTkComboBox(self,values=['Masculino','Feminino'],width=130,font=("Century Gohtic bold",14))
        gender_box.set("Masculino")
        gender_box.place(x=420, y=210)
        lb_obs = ctk.CTkLabel(self,text='Observações:',font=('Century Gothic bold',16),text_color=['#000','#fff']).place(x=50,y=240)
        obs_text_box = ctk.CTkTextbox(self,width=570,height=150,font=("arial",18),fg_color='transparent',border_color='#aaa',border_width=2)
        obs_text_box.place(x=50,y=270)

        bttn_submit= ctk.CTkButton(self,text="Enviar",width=120,command=submit,fg_color='#151',hover_color='#181').place(x=420,y=460)
        bttn_clear= ctk.CTkButton(self,text="Limpar",width=120,command=clear,fg_color='#611',hover_color='#911').place(x=560,y=460)


if __name__=="__main__":
    app = App()
    app.mainloop()
